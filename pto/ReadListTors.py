
# Maintenance V 1.0 (Автоопределение текущей даты, автоопределение текущей недели,
# вычисление номера недели для ТО, чтение любого количества графиков ТОР, вывод 
# запланированного ремонта на следующую неделю: для ПК - на экран и в Excel, для Android - на экран)

import openpyxl
import os.path
from termcolor import colored
from CalculateWeek import weekTO                            # Импорт номера недели для ТО из внешнего модуля

#print(wb.get_sheet_names())                                # Получение названий листов

path = 'L:/maintenance/charts/'                             # Открываем папку с файлами графиков
num_files = len([f for f in os.listdir(path)                # Получаем число 
                 if os.path.isfile(os.path.join(path, f))]) # файлов с графиками

#print(num_files)                                           # Раскомментировать в целях диагностики  
                                  
wbwrite = openpyxl.load_workbook('L:/maintenance/result/exel.xlsx')  # Выбираем книгу для записи
ws = wbwrite.active                                                  # Активируем ее
ws.delete_cols(1, 2)                                                 # Стираем две колонки
wbwrite.save('L:/maintenance/result/exel.xlsx')                      # Сохраняем книгу

for k in range(num_files):                                # Главный цикл по файлам в папке
    wb = openpyxl.load_workbook(f'L:/maintenance/charts/{k+1}.xlsx')     # Выбираем книгу 
    sheet = wb['Лист1']                                   # И рабочий лист
    m_row = sheet.max_row                                 # Получаем все заполненные строки листа
    wbwrite = openpyxl.load_workbook('L:/maintenance/result/exel.xlsx')  # Выбираем книгу для записи
    ws = wbwrite['Лист1']                                 # И лист для записи
    row = ws.max_row
    number = []                                           # Список номеров оборудования данного участка
    names = []                                            # Список наименований оборудования данного участка
    tors = []                                             # Общий список всех ТОР всего оборудования участка
    counter_i = 0                                         # Количество считанных ТОРов
    week_tors = []                                        # Перечень отдельных списков по ТОРам
    tor_list = ['ТОР-2', 'ТОР-3', 'ТОР-4', 'ТОР-5']

    '''
    for rowNumber in range(1, sheet.max_row + 1):         # Перебираем все строки 
        print(sheet.cell(row=rowNumber, column=6).value)  # в конкретных столбцах

    for colNumber in range(7, sheet.max_column + 1):      # Перебираем все столбцы 
        print(sheet.cell(column=colNumber, row=13).value) # в конкретных строках
    '''
    for i in range(1, m_row + 1):                          # Главный цикл по максимальному числу строк
        cell_obj = sheet.cell(row = i, column = 1)         # Читает первую строку первой колонки
        if isinstance(cell_obj.value, int):                # Проверяет, является ли полученное значение числом
            units = sheet.cell(row = i, column = 2)        # Читает вторую колонку в выбранной строке
            numberes = [cell_obj.value]                    # Получает значение из первой колонки в выбранной строке в виде списка
            name = [units.value]                           # Получает значение из второй колонки в выбранной строке в виде списка
            number.extend(numberes)                        # Добавляет значение в список номеров
            names.extend(name)                             # Добавляет значение в список имен
            #print(cell_obj.value)                         # Раскомментировать в целях диагностики
            for y in range(4):                                                  # Цикл считывает виды ремонта от ТОР-2 до ТОР-5
                for colNumber in range(7, sheet.max_column):                    # Перебираем все столбцы 
                    vals = [sheet.cell(column=colNumber, row=i + 2 + y).value]  # в конкретных строках
                    tors.extend(vals)                                           # Записываем полученные значения в общий список
                counter_i = counter_i + 1                                       # Счетчик считаных ТОРов

    #print(number)                                                              # Раскомментировать в целях диагностики
    #print(names)                                                               # Раскомментировать в целях диагностики
    #print(tors)                                                                # Раскомментировать в целях диагностики
    #print(counter_i)                                                           # Раскомментировать в целях диагностики

    countermash = 0                                                             # Динамический счетчик, для определения номера оборудования
    mash = []                                                                   # Номера строк с ТОР на заданную неделю                                                                
    numberTor = []                                                              # Список номеров оборудования, соответствует номерам строк ТОР
    counterTor = 0                                                              # Счетчик номера ТОР
    j = 0

    title = sheet.cell(row = 2, column = 2)                                     # Считывает заглавие листа
    #print()
    #print("\033[4m" + title.value + "\033[0m")                                 # Выводит заглавие листа
    #print()
    xxxx = [title.value]
    xxx = ['*******']
    xx = ['']
    #ws.append(xx)                                                              # Запись пробела
    #ws.append(xxx)                                                             # Запись звездочек
    #ws.append(xxxx)                                                            # Запись заглавия
    
    #print("\033[4mhello\033[0m")                                               # Подчеркнутый текст
    #print("\u0332".join("hello "))
    #print("\033[4m" + x + "\033[0m")

    for x in range(counter_i):                                                  # Главный цикл по количеству ТОРов
        if len(tors) > 0:                                                       # Проверка что общий список ТОРов еще не пуст
            tor = tors[0:48]                                                    # Срез первого по счету ТОРа
            del tors[0:48]                                                      # Его удаление из общего списка
            counterTor = counterTor + 1
            if x % 4 == 0:                                                      # Разделяет номера оборудования по 4 видам ремонта
                countermash = countermash + 1                                   # Повышает динамический счетчик
                counterTor = 0
            indices = [x + 1 for x in range(0, len(tor))
                        if tor[x]=="x" or tor[x]=="х"]                          # извлекаем номера недель, на которые запланирован ТОРы оборудования этого листа 
            #print(indices)
            for y in indices:                                                   # Цикл по полученным номерам недель одного ТОР 
                if y == weekTO:                                                 # Проверка на соответствие текущей неделе
                    j = j + 1
                    if j == 1:
                        print()
                        print(colored("\033[4m" + title.value + "\033[0m", 'green'))# Вывод заглавия листа
                        print()
                        ws.append(xx)                                               # Запись пробела
                        ws.append(xxx)                                              # Запись звездочек
                        ws.append(xxxx)                                             # Запись заглавия
                    mash.append(x + 1)                                              # Номера строк с нужными ТОР
                    numberTor.append(countermash)                                   # Соответствие номеров оборудования номерам 
                    ws.append([names[countermash-1][:]] + [tor_list[counterTor][:]])# Запись названия оборудования плюс вид ремонта
                    print((names[countermash-1][:]), (tor_list[counterTor][:]))     # ВЫвод на экран названия оборудования плюс вид ремонта 
                    print()

                              
    wbwrite.save('L:/maintenance/result/exel.xlsx') # Сохраняем данные
                                                                 
    #print(mash)                                                                # Раскомментировать в целях диагностики
    #print(countermash)                                                         # Раскомментировать в целях диагностики
    #print(numberTor)                                                           # Раскомментировать в целях диагностики
    #print(weekTO)                                                              # Раскомментировать в целях диагностики

    #indices = [x + 1 for x in range(0, len(item1_tor2)) if item1_tor2[x]=="x"] # Оригинал кода сопоставления значений

   
    